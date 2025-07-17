import os
from pathlib import Path, PureWindowsPath
import pandas as pd
from openpyxl import load_workbook

# 1)  Define the *parent* folders you want to search.

scan_roots = [
    PureWindowsPath(r"\\DEWDFG212.wdf.global.corp.sap\vul-scan\Reports\SPM\Scan_Results\priv"),
    PureWindowsPath(r"\\DEWDFG212.wdf.global.corp.sap\vul-scan\Reports\SPM-COM\Scan_Results\priv"),
    PureWindowsPath(r"\\DEWDFG212.wdf.global.corp.sap\vul-scan\Reports\SPM-ICM\Scan_Results\priv"),
]

# 2)  For each folder, grab the newest “*-PRIV-YYYYMMDD.xlsx” file.
#     (look at the modification time; change to .st_ctime for creation.)

input_files = []
for root in scan_roots:
    # *.xlsx files that match the usual naming scheme
    candidates = list(Path(root).glob("*-PRIV-*.xlsx"))
    if not candidates:
        print(f"No XLSX files found in {root}")
        continue

    newest = max(candidates, key=lambda p: p.stat().st_mtime)  # or st_ctime
    input_files.append(PureWindowsPath(newest))

# 3)  Read & combine columns A‑BJ from all the newest files.

dfs = []
for file in input_files:
    try:
        df = pd.read_excel(file, engine="openpyxl", usecols="A:BJ")
        # df["Source File"] = file  # uncomment if you want a provenance column
        dfs.append(df)
    except Exception as e:
        print(f"Error reading {file} – {e}")

if not dfs:
    raise RuntimeError("No valid Excel files were read — nothing to combine.")

final_df = pd.concat(dfs, ignore_index=True)

# 4)  Save the combined workbook.

output_file = "SPM-Combined-EVM-latest-automated.xlsx"
final_df.to_excel(output_file, index=False, engine="openpyxl")
print(f"Combined file saved as: {output_file}")

# 5) Apply formulas using openpyxl
wb = load_workbook(output_file)
ws = wb.active

# Headers for columns BK to BS
ws["BK1"] = "REBOOT"
ws["BL1"] = "DEVSERVER"
ws["BM1"] = "ServerType"
ws["BN1"] = "TEAM"
ws["BO1"] = "PRODSERVER"
ws["BP1"] = "OS Release"
ws["BQ1"] = "OS Release Patch"
ws["BR1"] = "OS Category"
ws["BS1"] = "mappedhostname"

for row in range(2, ws.max_row + 1):
    ws[f"BK{row}"] = (
        f'=((ISNUMBER(SEARCH("kernel", $G{row})) + ISNUMBER(SEARCH("glibc", $G{row})) + '
        f'ISNUMBER(SEARCH("ucode", $G{row})) + ISNUMBER(SEARCH("dbus", $G{row})) + '
        f'ISNUMBER(SEARCH("udev", $G{row})) + ISNUMBER(SEARCH("libudev", $G{row})) + '
        f'ISNUMBER(SEARCH("xen-kmp", $G{row})) + ISNUMBER(SEARCH("wicked", $G{row})) + '
        f'ISNUMBER(SEARCH("libwicked", $G{row})) + ISNUMBER(SEARCH("grub2", $G{row}))) > 0)'
    )
    ws[f"BL{row}"] = f'=ISNUMBER(SEARCH("-dev-",$C{row}))'

    ws[f"BM{row}"] = (
        f'=IF(ISNUMBER(SEARCH("lnd-",$C{row})),"LANDINGPAD",'
        f'IF(ISNUMBER(SEARCH("rpt-",$C{row})),"BO",'
        f'IF(ISNUMBER(SEARCH("ora-",$C{row})),"ORACLEDB",'
        f'IF(ISNUMBER(SEARCH("-hana-",$C{row})),"HANADB",'
        f'IF(ISNUMBER(SEARCH("pentaho",$C{row})),"PENTAHO",'
        f'IF(ISNUMBER(SEARCH("-wf-",$C{row})),"WORKFLOW",'
        f'IF(ISNUMBER(SEARCH("infa-",$C{row})),"INFORMATICA",'
        f'IF(ISNUMBER(SEARCH("infadb",$C{row})),"INFADB",'
        f'IF(ISNUMBER(SEARCH("app-comm",$C{row})),"COMMAPP",'
        f'IF(ISNUMBER(SEARCH("redis-",$C{row})),"REDIS",'
        f'IF(ISNUMBER(SEARCH("-lb-",$C{row})),"LOADBALANCER",'
        f'IF(ISNUMBER(SEARCH("kafka",$C{row})),"KAFKA",'
        f'IF(ISNUMBER(SEARCH("rabbitmq",$C{row})),"RABBITMQ",'
        f'IF(ISNUMBER(SEARCH("-cdl-",$C{row})),"CDL",'
        f'IF(ISNUMBER(SEARCH("app-pmpro-",$C{row})),"PMPRO","OTHER")'
        f'))))))))))))))'
    )
    ws[f"BN{row}"] = (
        f'=IF((ISNUMBER(SEARCH("RHEL", $G{row})) + ISNUMBER(SEARCH("SUSE", $G{row})) + '
        f'ISNUMBER(SEARCH("Sophos", $G{row})) + ISNUMBER(SEARCH("BMC", $G{row})) + '
        f'ISNUMBER(SEARCH("SNMP", $G{row})) + ISNUMBER(SEARCH("TLS", $G{row})) + '
        f'ISNUMBER(SEARCH("MYSQL", $G{row})) + ISNUMBER(SEARCH("SSH", $G{row})) + '
        f'ISNUMBER(SEARCH("SSL", $G{row})) + ISNUMBER(SEARCH("Unix Operating System Unsupported", $G{row}))), '
        f'"INFRA", "RM")'
    )
    ws[f"BO{row}"] = (
        f'=IF(ISERROR(VLOOKUP(LEFT($C{row}, FIND(".", $C{row})-1), '
        f'\'https://sap.sharepoint.com/teams/SPMVulnerabilityManagement/Shared Documents/General/EVM Latest Status/[active_server_list.xlsx]Sheet1\'!$A:$A, 1, FALSE)),"FALSE", "TRUE")'
    )
    ws[f"BP{row}"] = (
        f"=VLOOKUP('https://sap.sharepoint.com/Users/I354516/Desktop/EVM/EVM/SPM-Combined-EVM-NewTemplate.xlsx'!Table2[@[CCIR Object ID]], "
        f"'https://sap.sharepoint.com/Users/I354516/Downloads/SPM_VM_Report.xlsx'!Table1[#All],7,FALSE)"
    )
    ws[f"BQ{row}"] = (
        f"=VLOOKUP('https://sap.sharepoint.com/Users/I354516/Desktop/EVM/EVM/SPM-Combined-EVM-NewTemplate.xlsx'!Table2[@[CCIR Object ID]], "
        f"'https://sap.sharepoint.com/Users/I354516/Downloads/SPM_VM_Report.xlsx'!Table1[#All],8,FALSE)"
    )
    ws[f"BR{row}"] = (
        f"=VLOOKUP('https://sap.sharepoint.com/Users/I354516/Desktop/EVM/EVM/SPM-Combined-EVM-NewTemplate.xlsx'!Table2[@[CCIR Object ID]], "
        f"'https://sap.sharepoint.com/Users/I354516/Downloads/SPM_VM_Report.xlsx'!Table1[#All],5,FALSE)"
    )
    ws[f"BS{row}"] = (
        f"=VLOOKUP('https://sap.sharepoint.com/Users/I354516/Desktop/EVM/EVM/SPM-Combined-EVM-NewTemplate.xlsx'!Table2[@[CCIR Object ID]], "
        f"'https://sap.sharepoint.com/Users/I354516/Downloads/SPM_VM_Report.xlsx'!Table1[#All],3,FALSE)"
    )

# Final save locally in the GitHub runner's workspace
import os

# Get the workspace directory from environment variable (set by GitHub Actions)
workspace_dir = os.environ.get("GITHUB_WORKSPACE", os.getcwd())
output_file = os.path.join(workspace_dir, "SPM-Combined-EVM-latest-automated.xlsx")
wb.save(output_file)
print(f"Formulas added and final file saved as: {output_file}")

# Upload to SharePoint using Office365-REST-Python-Client
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential

# SharePoint site and folder details
sharepoint_site = "https://sap.sharepoint.com/sites/SPMVulnerabilityManagement"
target_folder = "/teams/SPMVulnerabilityManagement/Shared Documents/General/EVM Latest Status"

# Credentials should be set as environment variables in your GitHub Actions workflow
username = os.environ["SHAREPOINT_USERNAME"]
password = os.environ["SHAREPOINT_PASSWORD"]

# Read the file from the runner's workspace and upload
ctx = ClientContext(sharepoint_site).with_credentials(UserCredential(username, password))
with open(output_file, "rb") as content_file:
    file_content = content_file.read()
ctx.web.get_folder_by_server_relative_url(target_folder).upload_file("SPM-Combined-EVM-latest-automated.xlsx", file_content).execute_query()
print("File uploaded to SharePoint.")
