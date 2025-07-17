import pandas as pd
from openpyxl import load_workbook

# Step 1: Combine the Excel files (columns A–BJ)
input_files = [
    "C:\\Users\\I753931\\OneDrive - SAP SE\\Desktop\\Python\\Test\\SPM-COM-PRIV-20250402.xlsx",
    "C:\\Users\\I753931\\OneDrive - SAP SE\\Desktop\\Python\\Test\\SPM-ICM-PRIV-20250402.xlsx",
    "C:\\Users\\I753931\\OneDrive - SAP SE\\Desktop\\Python\\Test\\SPM-PRIV-20250402.xlsx"
]

combined_data = []
for file in input_files:
    df = pd.read_excel(file, engine='openpyxl', usecols="A:BJ")
    df['Source File'] = file
    combined_data.append(df)

final_df = pd.concat(combined_data, ignore_index=True)

# Save to intermediate file
output_file = "SPM-Combined-EVM 8.xlsx"
final_df.to_excel(output_file, index=False, engine='openpyxl')

# Step 2: Add formulas BK to BS using openpyxl
wb = load_workbook(output_file)
ws = wb.active

# Define formula templates (row will be formatted in)
for row in range(2, ws.max_row + 1):
    ws[f"BK{row}"] = f"""=((ISNUMBER(SEARCH("kernel", G{row})) + ISNUMBER(SEARCH("glibc", G{row})) + ISNUMBER(SEARCH("ucode", G{row})) + ISNUMBER(SEARCH("dbus", G{row})) + ISNUMBER(SEARCH("udev", G{row})) + ISNUMBER(SEARCH("libudev", G{row})) + ISNUMBER(SEARCH("xen-kmp", G{row})) + ISNUMBER(SEARCH("wicked", G{row})) + ISNUMBER(SEARCH("libwicked", G{row})) + ISNUMBER(SEARCH("grub2", G{row}))) > 0)"""
    
    ws[f"BL{row}"] = f'=ISNUMBER(SEARCH("-dev-", C{row}))'

    ws[f"BM{row}"] = (
        f'=IF(ISNUMBER(SEARCH("lnd-", C{row})), "LANDINGPAD", '
        f'IF(ISNUMBER(SEARCH("rpt-", C{row})), "BO", '
        f'IF(ISNUMBER(SEARCH("ora-", C{row})), "ORACLEDB", '
        f'IF(ISNUMBER(SEARCH("-hana-", C{row})), "HANADB", '
        f'IF(ISNUMBER(SEARCH("pentaho", C{row})), "PENTAHO", '
        f'IF(ISNUMBER(SEARCH("-wf-", C{row})), "WORKFLOW", '
        f'IF(ISNUMBER(SEARCH("infa-", C{row})), "INFORMATICA", '
        f'IF(ISNUMBER(SEARCH("infadb", C{row})), "INFADB", '
        f'IF(ISNUMBER(SEARCH("app-comm", C{row})), "COMMAPP", '
        f'IF(ISNUMBER(SEARCH("redis-", C{row})), "REDIS", '
        f'IF(ISNUMBER(SEARCH("-lb-", C{row})), "LOADBALANCER", '
        f'IF(ISNUMBER(SEARCH("kafka", C{row})), "KAFKA", '
        f'IF(ISNUMBER(SEARCH("rabbitmq", C{row})), "RABBITMQ", '
        f'IF(ISNUMBER(SEARCH("-cdl-", C{row})), "CDL", '
        f'IF(ISNUMBER(SEARCH("app-pmpro-", C{row})), "PMPRO", "OTHER"))))))))))))))'
    )

    ws[f"BN{row}"] = (
        f'=IF((ISNUMBER(SEARCH("RHEL", G{row})) + ISNUMBER(SEARCH("SUSE", G{row})) + '
        f'ISNUMBER(SEARCH("Sophos", G{row})) + ISNUMBER(SEARCH("BMC", G{row})) + '
        f'ISNUMBER(SEARCH("SNMP", G{row})) + ISNUMBER(SEARCH("TLS", G{row})) + '
        f'ISNUMBER(SEARCH("MYSQL", G{row})) + ISNUMBER(SEARCH("SSH", G{row})) + '
        f'ISNUMBER(SEARCH("SSL", G{row})) + ISNUMBER(SEARCH("Unix Operating System Unsupported", G{row}))), '
        f'"INFRA", "RM")'
    )

    ws[f"BO{row}"] = (
        f'=IF(ISERROR(VLOOKUP(LEFT(C{row}, FIND(".", C{row})-1), '
        f'\'[active_server_list.xlsx]Sheet1\'!A:A, 1, FALSE)), "FALSE", "TRUE")'
    )

    # BP to BS use Table lookup from external workbook
    ws[f"BP{row}"] = (
        f'=VLOOKUP([@[CCIR Object ID]], '
        f'\'[SPM_VM_Report.xlsx]Table1\'!A:Z, 7, FALSE)'
    )
    ws[f"BQ{row}"] = (
        f'=VLOOKUP([@[CCIR Object ID]], '
        f'\'[SPM_VM_Report.xlsx]Table1\'!A:Z, 8, FALSE)'
    )
    ws[f"BR{row}"] = (
        f'=VLOOKUP([@[CCIR Object ID]], '
        f'\'[SPM_VM_Report.xlsx]Table1\'!A:Z, 5, FALSE)'
    )
    ws[f"BS{row}"] = (
        f'=VLOOKUP([@[CCIR Object ID]], '
        f'\'[SPM_VM_Report.xlsx]Table1\'!A:Z, 3, FALSE)'
    )

# Save the final version
wb.save(output_file)
print(f"Combined file with formulas saved as: {output_file}")

 ws[f"BK{row}"] = f"""=((ISNUMBER(SEARCH("kernel", G{row})) + ISNUMBER(SEARCH("glibc", G{row})) + ISNUMBER(SEARCH("ucode", G{row})) + ISNUMBER(SEARCH("dbus", G{row})) + ISNUMBER(SEARCH("udev", G{row})) + ISNUMBER(SEARCH("libudev", G{row})) + ISNUMBER(SEARCH("xen-kmp", G{row})) + ISNUMBER(SEARCH("wicked", G{row})) + ISNUMBER(SEARCH("libwicked", G{row})) + ISNUMBER(SEARCH("grub2", G{row}))) > 0)"""
    
    ws[f"BL{row}"] = f'=ISNUMBER(SEARCH("-dev-", C{row}))'

    ws[f"BM{row}"] = (
        f'=IF(ISNUMBER(SEARCH("lnd-", C{row})), "LANDINGPAD", '
        f'IF(ISNUMBER(SEARCH("rpt-", C{row})), "BO", '
        f'IF(ISNUMBER(SEARCH("ora-", C{row})), "ORACLEDB", '
        f'IF(ISNUMBER(SEARCH("-hana-", C{row})), "HANADB", '
        f'IF(ISNUMBER(SEARCH("pentaho", C{row})), "PENTAHO", '
        f'IF(ISNUMBER(SEARCH("-wf-", C{row})), "WORKFLOW", '
        f'IF(ISNUMBER(SEARCH("infa-", C{row})), "INFORMATICA", '
        f'IF(ISNUMBER(SEARCH("infadb", C{row})), "INFADB", '
        f'IF(ISNUMBER(SEARCH("app-comm", C{row})), "COMMAPP", '
        f'IF(ISNUMBER(SEARCH("redis-", C{row})), "REDIS", '
        f'IF(ISNUMBER(SEARCH("-lb-", C{row})), "LOADBALANCER", '
        f'IF(ISNUMBER(SEARCH("kafka", C{row})), "KAFKA", '
        f'IF(ISNUMBER(SEARCH("rabbitmq", C{row})), "RABBITMQ", '
        f'IF(ISNUMBER(SEARCH("-cdl-", C{row})), "CDL", '
        f'IF(ISNUMBER(SEARCH("app-pmpro-", C{row})), "PMPRO", "OTHER"))))))))))))))'
    )

    ws[f"BN{row}"] = (
        f'=IF((ISNUMBER(SEARCH("RHEL", G{row})) + ISNUMBER(SEARCH("SUSE", G{row})) + '
        f'ISNUMBER(SEARCH("Sophos", G{row})) + ISNUMBER(SEARCH("BMC", G{row})) + '
        f'ISNUMBER(SEARCH("SNMP", G{row})) + ISNUMBER(SEARCH("TLS", G{row})) + '
        f'ISNUMBER(SEARCH("MYSQL", G{row})) + ISNUMBER(SEARCH("SSH", G{row})) + '
        f'ISNUMBER(SEARCH("SSL", G{row})) + ISNUMBER(SEARCH("Unix Operating System Unsupported", G{row}))), '
        f'"INFRA", "RM")'
    )

    ws[f"BO{row}"] = (
        f'=IF(ISERROR(VLOOKUP(LEFT(C{row}, FIND(".", C{row})-1), '
        f'\'[active_server_list.xlsx]Sheet1\'!A:A, 1, FALSE)), "FALSE", "TRUE")'
    )

    ws[f"BP{row}"] = (
        f'=VLOOKUP([@[CCIR Object ID]], '
        f'\'[SPM_VM_Report.xlsx]Table1\'!A:Z, 7, FALSE)'
    )
    ws[f"BQ{row}"] = (
        f'=VLOOKUP([@[CCIR Object ID]], '
        f'\'[SPM_VM_Report.xlsx]Table1\'!A:Z, 8, FALSE)'
    )
    ws[f"BR{row}"] = (
        f'=VLOOKUP([@[CCIR Object ID]], '
        f'\'[SPM_VM_Report.xlsx]Table1\'!A:Z, 5, FALSE)'
    )
    ws[f"BS{row}"] = (
        f'=VLOOKUP([@[CCIR Object ID]], '
        f'\'[SPM_VM_Report.xlsx]Table1\'!A:Z, 3, FALSE)'
    )